"""
통합 면접 스케줄링 GUI - 고급 스케줄링 + 디버깅 매트릭스
Port 8509 (advanced_gui.py) + Port 8510 (schedule_matrix_gui.py) 통합
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from typing import Dict, List, Tuple, Optional
from advanced_scheduler import AdvancedInterviewScheduler, TimeSlot
from improved_pdf_processor import process_pdf_file
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 페이지 설정
st.set_page_config(
    page_title="SNAAC 통합 면접 스케줄링 시스템",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS 스타일
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
        background-color: #f0f2f6;
        border-radius: 10px 10px 0 0;
    }
    .stTabs [aria-selected="true"] {
        background-color: #667eea;
        color: white;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        color: #155724;
        margin: 1rem 0;
    }
    .error-box {
        padding: 1rem;
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 5px;
        color: #721c24;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        color: #0c5460;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        background-color: #fff3cd;
        border: 1px solid #ffeeba;
        border-radius: 5px;
        color: #856404;
        margin: 1rem 0;
    }
    .metric-card {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
        margin: 0.5rem;
    }
    .team-card {
        background: #f8f9fa;
        padding: 0.8rem;
        border-radius: 8px;
        margin: 0.3rem 0;
        border-left: 4px solid #667eea;
    }
    .schedule-slot {
        background: white;
        padding: 0.5rem;
        margin: 0.2rem;
        border-radius: 5px;
        border: 1px solid #dee2e6;
    }
    .schedule-slot.group-a {
        border-left: 4px solid #28a745;
    }
    .schedule-slot.group-b {
        border-left: 4px solid #17a2b8;
    }
    .conflict-team {
        background: #fff3cd;
        padding: 0.3rem 0.6rem;
        border-radius: 4px;
        margin: 0.2rem;
        display: inline-block;
        border: 1px solid #ffc107;
    }
</style>
""", unsafe_allow_html=True)

# 헤더
st.markdown('<div class="main-header"><h1>🎯 SNAAC 통합 면접 스케줄링 시스템</h1><p>고급 스케줄링 + 디버깅 매트릭스 통합 버전</p></div>', unsafe_allow_html=True)

# 세션 상태 초기화
if 'teams' not in st.session_state:
    st.session_state.teams = {}
if 'schedule' not in st.session_state:
    st.session_state.schedule = {}
if 'scheduler' not in st.session_state:
    st.session_state.scheduler = None
if 'optimization_mode' not in st.session_state:
    st.session_state.optimization_mode = 'continuous'
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []

# 사이드바 설정
with st.sidebar:
    st.markdown("### ⚙️ 스케줄링 설정")
    
    # 최적화 모드 선택
    st.session_state.optimization_mode = st.selectbox(
        "최적화 모드",
        ['continuous', 'max_teams', 'balanced'],
        format_func=lambda x: {
            'continuous': '연속 배치 (면접관 편의 최적화)',
            'max_teams': '최대 팀 수 배치',
            'balanced': '균형 배치'
        }[x]
    )
    
    # A/B 그룹 설정
    use_ab_groups = st.checkbox("A/B 그룹 사용 (동시 2팀 면접)", value=True)
    
    # 시간대 설정
    st.markdown("### 📅 면접 시간 설정")
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("시작일", datetime(2024, 9, 12))
    with col2:
        end_date = st.date_input("종료일", datetime(2024, 9, 14))
    
    # 통계 표시
    st.markdown("### 📊 현재 상태")
    st.info(f"📁 업로드된 파일: {len(st.session_state.uploaded_files)}개")
    st.info(f"👥 등록된 팀: {len(st.session_state.teams)}개")
    if st.session_state.schedule and isinstance(st.session_state.schedule, dict):
        scheduled_count = sum(1 for slot in st.session_state.schedule.values() 
                            if isinstance(slot, dict) and (slot.get('group_a_team') or slot.get('group_b_team')))
        st.success(f"✅ 배치된 팀: {scheduled_count}개")

# 메인 탭
tabs = st.tabs([
    "📤 파일 업로드", 
    "👥 팀 정보 관리", 
    "🗓️ 스케줄링", 
    "📊 스케줄 매트릭스",
    "🔍 시간대별 분석",
    "📋 결과 확인",
    "💾 엑셀 다운로드",
    "🐛 디버깅"
])

# 탭 1: 파일 업로드
with tabs[0]:
    st.markdown("### 📤 PDF 파일 업로드")
    st.info("면접 신청서 PDF 파일을 업로드하세요. 여러 파일을 동시에 업로드할 수 있습니다.")
    
    uploaded_files = st.file_uploader(
        "PDF 파일 선택",
        type=['pdf'],
        accept_multiple_files=True,
        key="pdf_uploader_unified"
    )
    
    if uploaded_files:
        for file in uploaded_files:
            if file.name not in st.session_state.uploaded_files:
                st.session_state.uploaded_files.append(file.name)
        
        if st.button("📝 파일 처리 시작", key="process_files_unified"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for idx, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"처리 중: {uploaded_file.name}")
                progress_bar.progress((idx + 1) / len(uploaded_files))
                
                # 임시 파일로 저장
                temp_path = f"/tmp/{uploaded_file.name}"
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                
                # PDF 처리
                try:
                    result = process_pdf_file(temp_path)
                    team_name = result.get('팀명', f'팀_{idx+1}')
                    
                    if team_name and team_name != '미확인':
                        st.session_state.teams[team_name] = {
                            '대표자': result.get('대표자명', '미확인'),
                            '이메일': result.get('이메일', '미확인'),
                            '전화번호': result.get('전화번호', '미확인'),
                            '가능시간': result.get('면접 가능 시간', []),
                            '상세시간표': result.get('상세 시간표', {}),
                            '파일명': uploaded_file.name
                        }
                        st.success(f"✅ {team_name} 팀 정보 추출 완료")
                    else:
                        st.error(f"❌ {uploaded_file.name}: 팀 정보 추출 실패")
                    
                except Exception as e:
                    st.error(f"❌ {uploaded_file.name} 처리 오류: {str(e)}")
                
                # 임시 파일 삭제
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            
            status_text.text("✅ 모든 파일 처리 완료!")
            progress_bar.progress(1.0)

# 탭 2: 팀 정보 관리
with tabs[1]:
    st.markdown("### 👥 등록된 팀 정보")
    
    if st.session_state.teams:
        # 팀 정보 테이블
        team_data = []
        for team_name, info in st.session_state.teams.items():
            team_data.append({
                '팀명': team_name,
                '대표자': info['대표자'],
                '이메일': info['이메일'],
                '전화번호': info['전화번호'],
                '가능 시간 수': len(info['가능시간']),
                '파일명': info['파일명']
            })
        
        df_teams = pd.DataFrame(team_data)
        st.dataframe(df_teams, use_container_width=True)
        
        # 팀별 상세 정보
        st.markdown("### 📋 팀별 상세 정보")
        selected_team = st.selectbox("팀 선택", list(st.session_state.teams.keys()))
        
        if selected_team:
            team_info = st.session_state.teams[selected_team]
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"""
                <div class="team-card">
                    <h4>{selected_team}</h4>
                    <p>👤 대표자: {team_info['대표자']}</p>
                    <p>📧 이메일: {team_info['이메일']}</p>
                    <p>📞 전화: {team_info['전화번호']}</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown("**가능한 시간대:**")
                if team_info['가능시간']:
                    for time in team_info['가능시간'][:10]:  # 처음 10개만 표시
                        st.write(f"• {time}")
                    if len(team_info['가능시간']) > 10:
                        st.write(f"... 외 {len(team_info['가능시간'])-10}개")
                else:
                    st.warning("가능한 시간대가 없습니다.")
    else:
        st.info("아직 등록된 팀이 없습니다. PDF 파일을 업로드해주세요.")

# 탭 3: 스케줄링
with tabs[2]:
    st.markdown("### 🗓️ 면접 스케줄 생성")
    
    if st.session_state.teams:
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🚀 자동 스케줄링 시작", key="auto_schedule_unified"):
                with st.spinner("스케줄 생성 중..."):
                    # 스케줄러 생성
                    scheduler = AdvancedInterviewScheduler()
                    
                    # 팀 정보 추가
                    for team_name, info in st.session_state.teams.items():
                        scheduler.add_team(team_name, info['가능시간'])
                    
                    # 스케줄링 실행
                    if st.session_state.optimization_mode == 'continuous':
                        result = scheduler.schedule_interviews_continuous()
                    elif st.session_state.optimization_mode == 'max_teams':
                        result = scheduler.schedule_interviews_max_teams()
                    else:
                        result = scheduler.schedule_interviews_balanced()
                    
                    # 결과가 튜플인지 확인
                    if isinstance(result, tuple) and len(result) == 2:
                        schedule, stats = result
                    else:
                        schedule = result
                        # 통계 직접 계산 - 배치된 팀 찾기
                        assigned_teams_set = set()
                        for slot in scheduler.time_slots.values():
                            if slot.group_a_team:
                                assigned_teams_set.add(slot.group_a_team)
                            if slot.group_b_team:
                                assigned_teams_set.add(slot.group_b_team)
                        
                        # scheduler.teams가 Team 객체의 dict인지 list인지 확인
                        all_team_names = []
                        if isinstance(scheduler.teams, dict):
                            all_team_names = list(scheduler.teams.keys())
                        elif isinstance(scheduler.teams, list):
                            # Team 객체 리스트인 경우
                            all_team_names = [team.name if hasattr(team, 'name') else str(team) 
                                            for team in scheduler.teams]
                        
                        unassigned_teams = [team for team in all_team_names 
                                          if team not in assigned_teams_set]
                        
                        stats = {
                            'assigned_teams': len(assigned_teams_set),
                            'unassigned_teams': len(unassigned_teams),
                            'used_slots': sum(1 for slot in scheduler.time_slots.values() 
                                            if slot.group_a_team or slot.group_b_team),
                            'total_slots': len(scheduler.time_slots),
                            'unassigned_list': unassigned_teams,
                            'conflicts': {}
                        }
                    
                    # schedule 형식 변환 (TimeSlot 객체를 dict로)
                    if schedule:
                        converted_schedule = {}
                        for slot_key, slot_value in schedule.items():
                            if hasattr(slot_value, '__dict__'):  # TimeSlot 객체인 경우
                                converted_schedule[slot_key] = {
                                    'group_a_team': slot_value.group_a_team,
                                    'group_b_team': slot_value.group_b_team
                                }
                            else:
                                converted_schedule[slot_key] = slot_value
                        st.session_state.schedule = converted_schedule
                    else:
                        st.session_state.schedule = {}
                    
                    st.session_state.scheduler = scheduler
                    
                    # 결과 표시
                    st.success(f"✅ 스케줄링 완료!")
                    st.write(f"• 배치된 팀: {stats['assigned_teams']}개")
                    st.write(f"• 미배치 팀: {stats['unassigned_teams']}개")
                    st.write(f"• 사용된 시간대: {stats['used_slots']}개")
                    st.write(f"• 총 시간대: {stats['total_slots']}개")
                    
                    # 미배치 팀 정보
                    if stats['unassigned_list']:
                        st.warning(f"⚠️ 미배치 팀: {', '.join(stats['unassigned_list'])}")
                        
                        # 충돌 정보 표시
                        if stats.get('conflicts'):
                            st.markdown("### 🔍 미배치 원인 분석")
                            for team, conflict_info in stats['conflicts'].items():
                                st.markdown(f"**{team}**")
                                st.write(f"• 가능 시간: {conflict_info['available_count']}개")
                                if conflict_info['conflicting_teams']:
                                    st.write(f"• 충돌 팀: {', '.join(list(conflict_info['conflicting_teams'])[:5])}")
        
        with col2:
            if st.button("🔄 스케줄 초기화", key="reset_schedule_unified"):
                st.session_state.schedule = {}
                st.session_state.scheduler = None
                st.success("스케줄이 초기화되었습니다.")
        
        with col3:
            if st.button("📊 최적화 분석", key="analyze_optimization_unified"):
                if st.session_state.scheduler:
                    gaps = st.session_state.scheduler.calculate_gaps()
                    st.metric("빈 시간대", f"{gaps}개")
    else:
        st.info("먼저 팀 정보를 업로드해주세요.")

# 탭 4: 스케줄 매트릭스 (디버깅용)
with tabs[3]:
    st.markdown("### 📊 스케줄 매트릭스 뷰")
    st.info("각 팀이 선택한 시간대를 한눈에 확인할 수 있습니다.")
    
    if st.session_state.teams:
        # 시간대 정의
        dates = ["9/12 (금)", "9/13 (토)", "9/14 (일)"]
        time_slots_by_date = {
            "9/12 (금)": ["19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"],
            "9/13 (토)": ["10:00-10:45", "10:45-11:30", "11:30-12:15", "12:15-13:00",
                         "13:00-13:45", "13:45-14:30", "14:30-15:15", "15:15-16:00",
                         "16:00-16:45", "16:45-17:30", "17:30-18:15", "18:15-19:00",
                         "19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"],
            "9/14 (일)": ["10:00-10:45", "10:45-11:30", "11:30-12:15", "12:15-13:00",
                         "13:00-13:45", "13:45-14:30", "14:30-15:15", "15:15-16:00",
                         "16:00-16:45", "16:45-17:30", "17:30-18:15", "18:15-19:00",
                         "19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"]
        }
        
        # 날짜별 탭
        date_tabs = st.tabs(dates)
        
        for date_idx, date in enumerate(dates):
            with date_tabs[date_idx]:
                date_key = date.split(' ')[0]  # "9/12"
                time_slots = time_slots_by_date[date]
                
                # 매트릭스 데이터 생성
                matrix_data = []
                team_names = list(st.session_state.teams.keys())
                
                for team_name in team_names:
                    row = {'팀명': team_name}
                    team_info = st.session_state.teams[team_name]
                    
                    for time_slot in time_slots:
                        # 해당 시간대가 가능한지 확인
                        full_slot = f"{date_key} {time_slot}"
                        if full_slot in team_info['가능시간']:
                            row[time_slot] = '⭕'
                        else:
                            row[time_slot] = ''
                    
                    matrix_data.append(row)
                
                # 데이터프레임 생성
                df_matrix = pd.DataFrame(matrix_data)
                
                # 스타일 적용
                def highlight_available(val):
                    if val == '⭕':
                        return 'background-color: #d4edda; color: #155724; font-weight: bold;'
                    return ''
                
                styled_df = df_matrix.style.applymap(highlight_available)
                st.dataframe(styled_df, use_container_width=True, height=400)
                
                # 통계
                st.markdown(f"### 📊 {date} 통계")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    total_teams = len(team_names)
                    st.metric("총 팀 수", f"{total_teams}팀")
                
                with col2:
                    # 각 시간대별 가능 팀 수 계산
                    avg_teams_per_slot = 0
                    for time_slot in time_slots:
                        count = sum(1 for row in matrix_data if row[time_slot] == '⭕')
                        avg_teams_per_slot += count
                    avg_teams_per_slot = avg_teams_per_slot / len(time_slots) if time_slots else 0
                    st.metric("평균 가능 팀/시간", f"{avg_teams_per_slot:.1f}팀")
                
                with col3:
                    total_slots = len(time_slots)
                    st.metric("총 시간대", f"{total_slots}개")

# 탭 5: 시간대별 분석
with tabs[4]:
    st.markdown("### 🔍 시간대별 가능 팀 분석")
    
    if st.session_state.teams:
        # 모든 시간대 수집
        all_slots = []
        for date, slots in time_slots_by_date.items():
            date_key = date.split(' ')[0]
            for slot in slots:
                all_slots.append(f"{date_key} {slot}")
        
        # 시간대별 가능 팀 계산
        slot_analysis = []
        for slot in all_slots:
            available_teams = []
            for team_name, team_info in st.session_state.teams.items():
                if slot in team_info['가능시간']:
                    available_teams.append(team_name)
            
            slot_analysis.append({
                '시간대': slot,
                '가능 팀 수': len(available_teams),
                '가능 팀': ', '.join(available_teams[:5]) + ('...' if len(available_teams) > 5 else '')
            })
        
        # 데이터프레임 생성
        df_analysis = pd.DataFrame(slot_analysis)
        df_analysis = df_analysis.sort_values('가능 팀 수', ascending=False)
        
        # 컬러 코딩
        def color_by_count(val):
            if isinstance(val, int):
                if val == 0:
                    return 'background-color: #f8d7da; color: #721c24;'
                elif val <= 2:
                    return 'background-color: #fff3cd; color: #856404;'
                elif val <= 5:
                    return 'background-color: #d1ecf1; color: #0c5460;'
                else:
                    return 'background-color: #d4edda; color: #155724;'
            return ''
        
        styled_analysis = df_analysis.style.applymap(color_by_count, subset=['가능 팀 수'])
        st.dataframe(styled_analysis, use_container_width=True, height=600)
        
        # 차트
        st.markdown("### 📈 시간대별 경쟁률")
        chart_data = df_analysis.set_index('시간대')['가능 팀 수']
        st.bar_chart(chart_data)

# 탭 6: 결과 확인
with tabs[5]:
    st.markdown("### 📋 최종 스케줄 결과")
    
    if st.session_state.schedule:
        # 전체 통계
        total_slots = len(st.session_state.schedule)
        total_a_teams = sum(1 for slot in st.session_state.schedule.values() if slot.get('group_a_team') and slot.get('group_a_team') != '-')
        total_b_teams = sum(1 for slot in st.session_state.schedule.values() if slot.get('group_b_team') and slot.get('group_b_team') != '-')
        total_teams_scheduled = total_a_teams + total_b_teams
        
        # 통계 표시
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📊 총 시간대", f"{total_slots}개")
        with col2:
            st.metric("🅰️ A그룹 배치", f"{total_a_teams}팀")
        with col3:
            st.metric("🅱️ B그룹 배치", f"{total_b_teams}팀")
        with col4:
            st.metric("✅ 총 배치", f"{total_teams_scheduled}팀")
        
        st.markdown("---")
        
        # 날짜별 스케줄 정리
        schedule_by_date = {
            "9/12": [],
            "9/13": [],
            "9/14": []
        }
        
        for slot_key, slot_info in st.session_state.schedule.items():
            parts = slot_key.split(' ')
            if len(parts) >= 2:
                date = parts[0]
                time = ' '.join(parts[1:])
            else:
                date = slot_key
                time = ""
            
            entry = {
                '시간': time,
                'A그룹': slot_info.get('group_a_team', '-'),
                'B그룹': slot_info.get('group_b_team', '-')
            }
            
            if date in schedule_by_date:
                schedule_by_date[date].append(entry)
        
        # 날짜별 표시 (개선된 버전)
        date_names = {"9/12": "9월 12일 (금)", "9/13": "9월 13일 (토)", "9/14": "9월 14일 (일)"}
        
        for date, entries in schedule_by_date.items():
            if entries:
                st.markdown(f"### 📅 {date_names.get(date, date)}")
                
                # 시간순 정렬
                entries = sorted(entries, key=lambda x: x['시간'])
                
                # 스타일링된 표시
                for i, entry in enumerate(entries):
                    with st.container():
                        # 시간대별 카드
                        time_slot = entry['시간']
                        a_team = entry['A그룹']
                        b_team = entry['B그룹']
                        
                        # 시간대 헤더
                        st.markdown(f"**⏰ {time_slot}**")
                        
                        # A조/B조 나란히 표시
                        col_a, col_b = st.columns(2)
                        
                        with col_a:
                            if a_team and a_team != '-':
                                st.markdown(f"""
                                <div class="team-card" style="background: #e8f5e8; border-left: 4px solid #28a745;">
                                    <h4 style="margin: 0; color: #155724;">🅰️ A그룹</h4>
                                    <p style="margin: 5px 0; font-weight: bold;">{a_team}</p>
                                    <small style="color: #666;">면접실 A</small>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                st.markdown(f"""
                                <div class="team-card" style="background: #f8f9fa; border-left: 4px solid #dee2e6;">
                                    <h4 style="margin: 0; color: #6c757d;">🅰️ A그룹</h4>
                                    <p style="margin: 5px 0; color: #999;">빈 자리</p>
                                    <small style="color: #666;">면접실 A</small>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        with col_b:
                            if b_team and b_team != '-':
                                st.markdown(f"""
                                <div class="team-card" style="background: #e3f2fd; border-left: 4px solid #2196f3;">
                                    <h4 style="margin: 0; color: #0d47a1;">🅱️ B그룹</h4>
                                    <p style="margin: 5px 0; font-weight: bold;">{b_team}</p>
                                    <small style="color: #666;">면접실 B</small>
                                </div>
                                """, unsafe_allow_html=True)
                            else:
                                st.markdown(f"""
                                <div class="team-card" style="background: #f8f9fa; border-left: 4px solid #dee2e6;">
                                    <h4 style="margin: 0; color: #6c757d;">🅱️ B그룹</h4>
                                    <p style="margin: 5px 0; color: #999;">빈 자리</p>
                                    <small style="color: #666;">면접실 B</small>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        if i < len(entries) - 1:
                            st.markdown("---")
                
                # 날짜별 통계
                date_a_teams = sum(1 for entry in entries if entry['A그룹'] and entry['A그룹'] != '-')
                date_b_teams = sum(1 for entry in entries if entry['B그룹'] and entry['B그룹'] != '-')
                date_total = date_a_teams + date_b_teams
                date_slots = len(entries)
                utilization = (date_total / (date_slots * 2) * 100) if date_slots > 0 else 0
                
                st.markdown(f"""
                <div style="background: #f1f3f4; padding: 1rem; border-radius: 8px; margin: 1rem 0;">
                    <strong>📊 {date_names.get(date, date)} 통계</strong><br>
                    • 총 시간대: {date_slots}개 • A그룹: {date_a_teams}팀 • B그룹: {date_b_teams}팀 • 활용률: {utilization:.1f}%
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("---")
                
            else:
                st.info(f"{date_names.get(date, date)}에 배치된 팀이 없습니다.")
                
        # 빈 시간대 분석
        if hasattr(st.session_state, 'scheduler') and st.session_state.scheduler:
            gaps = st.session_state.scheduler.calculate_gaps()
            if gaps:
                st.markdown("### ⚠️ 빈 시간대 분석")
                gaps_df = pd.DataFrame(gaps)
                st.dataframe(gaps_df, use_container_width=True)
    else:
        st.info("아직 생성된 스케줄이 없습니다. '스케줄링' 탭에서 스케줄을 생성해주세요.")

# 탭 7: 엑셀 다운로드
with tabs[6]:
    st.markdown("### 💾 엑셀 파일 다운로드")
    
    if st.session_state.schedule:
        if st.button("📥 엑셀 파일 생성", key="create_excel_unified"):
            # 엑셀 파일 생성
            from io import BytesIO
            
            output = BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "면접 스케줄"
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # 헤더 작성
            headers = ["날짜", "시간", "A그룹 팀명", "A그룹 대표자", "B그룹 팀명", "B그룹 대표자"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터 작성
            row_num = 2
            for slot_key, slot_info in sorted(st.session_state.schedule.items()):
                # slot_key 형식 확인 및 파싱
                parts = slot_key.split(' ')
                if len(parts) >= 2:
                    date = parts[0]
                    time = ' '.join(parts[1:])
                else:
                    date = slot_key
                    time = ""
                
                ws.cell(row=row_num, column=1, value=date)
                ws.cell(row=row_num, column=2, value=time)
                
                # A그룹
                a_team = slot_info.get('group_a_team', '')
                ws.cell(row=row_num, column=3, value=a_team)
                if a_team and a_team in st.session_state.teams:
                    ws.cell(row=row_num, column=4, value=st.session_state.teams[a_team]['대표자'])
                
                # B그룹
                b_team = slot_info.get('group_b_team', '')
                ws.cell(row=row_num, column=5, value=b_team)
                if b_team and b_team in st.session_state.teams:
                    ws.cell(row=row_num, column=6, value=st.session_state.teams[b_team]['대표자'])
                
                row_num += 1
            
            # 열 너비 조정
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 20
            
            # 파일 저장
            wb.save(output)
            output.seek(0)
            
            # 다운로드 버튼
            st.download_button(
                label="📥 엑셀 파일 다운로드",
                data=output.getvalue(),
                file_name=f"SNAAC_면접스케줄_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ 엑셀 파일이 준비되었습니다!")
    else:
        st.info("먼저 스케줄을 생성해주세요.")

# 탭 8: 디버깅
with tabs[7]:
    st.markdown("### 🐛 디버깅 도구")
    
    # 팀별 상세 정보 확인
    st.markdown("#### 🔍 팀별 상세 정보 확인")
    if st.session_state.teams:
        debug_team = st.selectbox("디버깅할 팀 선택", list(st.session_state.teams.keys()), key="debug_team_select")
        
        if debug_team:
            team_info = st.session_state.teams[debug_team]
            
            # JSON 형태로 상세 정보 표시
            st.json({
                '팀명': debug_team,
                '대표자': team_info['대표자'],
                '이메일': team_info['이메일'],
                '전화번호': team_info['전화번호'],
                '가능시간_개수': len(team_info['가능시간']),
                '가능시간_목록': team_info['가능시간'],
                '파일명': team_info['파일명']
            })
            
            # 상세 시간표가 있는 경우
            if team_info.get('상세시간표'):
                st.markdown("##### 📅 상세 시간표")
                for date, slots in team_info['상세시간표'].items():
                    st.write(f"**{date}:**")
                    for time, available in slots:
                        status = "✅ 가능" if available else "❌ 불가"
                        st.write(f"  • {time}: {status}")
    
    # 스케줄러 상태 확인
    st.markdown("#### 🔧 스케줄러 상태")
    if st.session_state.scheduler:
        scheduler = st.session_state.scheduler
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("총 팀 수", len(scheduler.teams))
            st.metric("총 시간대", len(scheduler.time_slots))
        
        with col2:
            assigned = sum(1 for slot in scheduler.time_slots.values() 
                         if slot.group_a_team or slot.group_b_team)
            st.metric("배치된 슬롯", assigned)
            
            # 배치된 팀 찾기
            assigned_teams_in_scheduler = set()
            for slot in scheduler.time_slots.values():
                if slot.group_a_team:
                    assigned_teams_in_scheduler.add(slot.group_a_team)
                if slot.group_b_team:
                    assigned_teams_in_scheduler.add(slot.group_b_team)
            
            # scheduler.teams 타입 확인
            all_team_names_debug = []
            if isinstance(scheduler.teams, dict):
                all_team_names_debug = list(scheduler.teams.keys())
            elif isinstance(scheduler.teams, list):
                all_team_names_debug = [team.name if hasattr(team, 'name') else str(team) 
                                       for team in scheduler.teams]
            
            unassigned = [team for team in all_team_names_debug 
                         if team not in assigned_teams_in_scheduler]
            st.metric("미배치 팀", len(unassigned))
        
        # 미배치 팀 상세
        if unassigned:
            st.markdown("##### ⚠️ 미배치 팀 상세")
            for team in unassigned:
                team_info = st.session_state.teams.get(team, {})
                st.write(f"• **{team}**: {len(team_info.get('가능시간', []))}개 시간 가능")
    
    # 시스템 로그
    st.markdown("#### 📝 시스템 로그")
    if st.button("시스템 정보 새로고침", key="refresh_debug"):
        st.rerun()
    
    # 세션 상태 덤프
    with st.expander("세션 상태 전체 보기"):
        st.json({
            'teams_count': len(st.session_state.teams),
            'schedule_count': len(st.session_state.schedule),
            'uploaded_files': st.session_state.uploaded_files,
            'optimization_mode': st.session_state.optimization_mode
        })