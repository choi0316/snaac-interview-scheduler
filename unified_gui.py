"""
통합 면접 스케줄링 GUI - 고급 스케줄링 + 디버깅 매트릭스
Port 8509 (advanced_gui.py) + Port 8510 (schedule_matrix_gui.py) 통합
calculate_gaps 메서드 누락 오류 수정 및 A/B 조 스케줄 표시 개선
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from typing import Dict, List, Tuple, Optional
from advanced_scheduler import AdvancedInterviewScheduler, TimeSlot
from improved_pdf_processor import process_pdf_file
from team_editor_component import render_team_editor, render_manual_team_adder
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

# 페이지 설정
st.set_page_config(
    page_title="SNAAC 통합 면접 스케줄링 시스템",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS 스타일
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-left: 20px;
        padding-right: 20px;
        background-color: #f0f2f6;
        border-radius: 10px 10px 0 0;
    }
    .stTabs [aria-selected="true"] {
        background-color: #667eea;
        color: white;
    }
    .team-card {
        background: #f8f9fa;
        padding: 0.8rem;
        border-radius: 8px;
        margin: 0.3rem 0;
        border-left: 4px solid #667eea;
    }
</style>
""", unsafe_allow_html=True)

# 세션 상태 초기화
if 'teams' not in st.session_state:
    st.session_state.teams = {}
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []
if 'schedule' not in st.session_state:
    st.session_state.schedule = None
if 'scheduler' not in st.session_state:
    st.session_state.scheduler = None
if 'stats' not in st.session_state:
    st.session_state.stats = {}

# 헤더
st.markdown("""
<div class="main-header">
    <h1>🏢 SNAAC 통합 면접 스케줄링 시스템</h1>
    <p>AI 기반 최적 스케줄링 + 실시간 매트릭스 분석</p>
</div>
""", unsafe_allow_html=True)

# 사이드바
st.sidebar.markdown("### ⚙️ 스케줄링 설정")
optimization_mode = st.sidebar.selectbox(
    "최적화 모드",
    ["continuous", "max_teams", "interviewer_friendly"],
    format_func=lambda x: {
        "continuous": "🔄 연속 스케줄링 (간격 최소화)",
        "max_teams": "👥 최대 팀 배치",
        "interviewer_friendly": "🤝 면접관 친화적 (짝 맞춰 배치)"
    }[x]
)

# 메인 탭
tabs = st.tabs([
    "📤 파일 업로드", 
    "👥 팀 관리", 
    "🗓️ 스케줄링", 
    "📋 A/B조 스케줄",
    "📊 매트릭스 뷰",
    "💾 엑셀 다운로드"
])

# 탭 1: 파일 업로드
with tabs[0]:
    st.markdown("### 📤 PDF 파일 업로드")
    st.info("면접 신청서 PDF 파일을 업로드하세요.")
    
    uploaded_files = st.file_uploader(
        "PDF 파일 선택",
        type=['pdf'],
        accept_multiple_files=True,
        key="pdf_uploader"
    )
    
    if uploaded_files:
        if st.button("📝 파일 처리 시작", key="process_files"):
            progress_bar = st.progress(0)
            
            for idx, uploaded_file in enumerate(uploaded_files):
                progress_bar.progress((idx + 1) / len(uploaded_files))
                
                # 임시 파일로 저장
                temp_path = f"/tmp/{uploaded_file.name}"
                with open(temp_path, "wb") as f:
                    f.write(uploaded_file.read())
                
                try:
                    # PDF 처리
                    team_info = process_pdf_file(temp_path)
                    if team_info and team_info.get('팀명'):
                        team_name = team_info['팀명']
                        # PDF 파일명 추가
                        team_info['파일명'] = uploaded_file.name
                        st.session_state.teams[team_name] = team_info
                        st.success(f"✅ {team_name} 처리 완료")
                    else:
                        st.error(f"❌ {uploaded_file.name} 처리 실패")
                except Exception as e:
                    st.error(f"❌ {uploaded_file.name} 처리 중 오류: {e}")
                finally:
                    # 임시 파일 삭제
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
            
            progress_bar.progress(1.0)
            st.success(f"🎉 총 {len(st.session_state.teams)}개 팀 처리 완료!")
            st.rerun()

# 탭 2: 팀 관리
with tabs[1]:
    st.markdown("### 👥 등록된 팀 정보")
    
    if st.session_state.teams:
        # 팀 목록과 상세정보를 나란히 표시 (반반 비율)
        col_list, col_detail = st.columns([1, 1])
        
        with col_list:
            st.markdown("**📋 팀 목록**")
            team_data = []
            for team_name, info in st.session_state.teams.items():
                team_data.append({
                    '팀명': team_name,
                    '대표자': info.get('대표자명', 'N/A'),
                    '파일명': info.get('파일명', 'N/A'),
                    '가능 시간수': len(info.get('가능시간', []))
                })
            
            df_teams = pd.DataFrame(team_data)
            st.dataframe(df_teams, use_container_width=True, hide_index=True)
            
            # 팀 선택
            st.markdown("---")
            selected_team = st.selectbox(
                "🔍 상세정보 볼 팀 선택",
                options=["선택하세요..."] + list(st.session_state.teams.keys()),
                key="team_selector"
            )
        
        with col_detail:
            st.markdown("**📄 팀 상세정보**")
            
            if selected_team and selected_team != "선택하세요...":
                team_info = st.session_state.teams[selected_team]
                
                # 편집 모드 체크
                edit_mode_key = f"edit_mode_{selected_team}"
                
                # 편집 버튼을 상단에 배치
                col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 3])
                with col_btn1:
                    if st.button("✏️ 수정", key=f"edit_{selected_team}_btn"):
                        st.session_state[edit_mode_key] = not st.session_state.get(edit_mode_key, False)
                        st.rerun()
                with col_btn2:
                    if st.button("🗑️ 삭제", key=f"delete_{selected_team}_btn"):
                        del st.session_state.teams[selected_team]
                        st.success(f"✅ {selected_team} 팀이 삭제되었습니다.")
                        st.rerun()
                
                # 편집 모드인 경우
                if st.session_state.get(edit_mode_key, False):
                    with st.expander("✏️ 팀 정보 수정", expanded=True):
                        # 팀 정보 편집 UI 렌더링
                        updated_info = render_team_editor(selected_team, team_info)
                        
                        # 저장 및 취소 버튼
                        col_save, col_cancel, col_empty = st.columns([1, 1, 3])
                        with col_save:
                            if st.button("💾 저장", key=f"save_{selected_team}"):
                                # 업데이트된 정보를 세션에 저장
                                st.session_state.teams[selected_team] = updated_info
                                st.session_state[edit_mode_key] = False
                                st.success(f"✅ {selected_team} 팀 정보가 업데이트되었습니다.")
                                st.rerun()
                        
                        with col_cancel:
                            if st.button("❌ 취소", key=f"cancel_{selected_team}"):
                                st.session_state[edit_mode_key] = False
                                st.rerun()
                else:
                    # 상세정보 카드 (읽기 전용)
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                               color: white; padding: 1.5rem; border-radius: 10px; margin-bottom: 1rem;">
                        <h3 style="margin: 0 0 1rem 0;">🏢 {selected_team}</h3>
                        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 1rem;">
                            <div>
                                <strong>👤 대표자:</strong><br>{team_info.get('대표자명', '정보 없음')}
                            </div>
                            <div>
                                <strong>📧 이메일:</strong><br>{team_info.get('이메일', '정보 없음')}
                            </div>
                            <div>
                                <strong>📞 전화번호:</strong><br>{team_info.get('전화번호', '정보 없음')}
                            </div>
                            <div>
                                <strong>⏰ 가능 시간:</strong><br>{len(team_info.get('가능시간', []))}개 시간대
                            </div>
                            <div style="grid-column: 1 / -1; margin-top: 0.5rem;">
                                <strong>📄 파일명:</strong><br>{team_info.get('파일명', '정보 없음')}
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # 가능 시간 목록
                    if team_info.get('가능시간'):
                        st.markdown("**📅 면접 가능 시간 목록**")
                        
                        # 날짜별로 정리
                        time_by_date = {"9/12": [], "9/13": [], "9/14": []}
                        for time_slot in team_info.get('가능시간', []):
                            for date in time_by_date.keys():
                                if time_slot.startswith(date):
                                    time_by_date[date].append(time_slot.replace(f"{date} ", ""))
                        
                        # 날짜별 표시
                        date_names = {"9/12": "9월 12일 (금)", "9/13": "9월 13일 (토)", "9/14": "9월 14일 (일)"}
                        for date, times in time_by_date.items():
                            if times:
                                with st.expander(f"📅 {date_names[date]} ({len(times)}개 시간)"):
                                    for time in sorted(times):
                                        st.write(f"• {time}")
                    else:
                        st.info("면접 가능 시간 정보가 없습니다. '수정' 버튼을 눌러 시간대를 추가할 수 있습니다.")
            else:
                st.info("좌측에서 팀을 선택하면 상세정보가 표시됩니다.")
        
        # 수동 팀 추가 섹션 (하단에 배치)
        st.markdown("---")
        with st.expander("➕ 새 팀 수동 추가", expanded=False):
            result = render_manual_team_adder()
            if result:
                team_name, team_info = result
                if team_name not in st.session_state.teams:
                    st.session_state.teams[team_name] = team_info
                    st.success(f"✅ {team_name} 팀이 추가되었습니다!")
                    st.rerun()
                else:
                    st.error(f"이미 {team_name}이라는 팀이 존재합니다.")
        
        # 통계 (하단에 표시)
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("총 팀 수", len(st.session_state.teams))
        with col2:
            avg_slots = sum(len(info.get('가능시간', [])) for info in st.session_state.teams.values()) / len(st.session_state.teams)
            st.metric("평균 가능 시간", f"{avg_slots:.1f}개")
        with col3:
            st.metric("업로드 파일", len(st.session_state.uploaded_files))
    else:
        st.info("아직 업로드된 팀이 없습니다.")
        
        # 수동 팀 추가 섹션
        with st.expander("➕ 새 팀 수동 추가", expanded=True):
            result = render_manual_team_adder()
            if result:
                team_name, team_info = result
                if team_name not in st.session_state.teams:
                    st.session_state.teams[team_name] = team_info
                    st.success(f"✅ {team_name} 팀이 추가되었습니다!")
                    st.rerun()
                else:
                    st.error(f"이미 {team_name}이라는 팀이 존재합니다.")

# 탭 3: 스케줄링
with tabs[2]:
    st.markdown("### 🗓️ 스케줄링 실행")
    
    if st.session_state.teams:
        st.info(f"현재 {len(st.session_state.teams)}개 팀이 등록되어 있습니다.")
        
        if st.button("🚀 스케줄링 시작", key="start_scheduling"):
            with st.spinner("스케줄링 진행 중..."):
                try:
                    # 스케줄러 초기화
                    scheduler = AdvancedInterviewScheduler()
                    
                    # 팀 추가
                    for team_name, info in st.session_state.teams.items():
                        scheduler.add_team(team_name, info['가능시간'])
                    
                    # 스케줄링 실행
                    if optimization_mode == 'continuous':
                        result = scheduler.schedule_interviews_continuous()
                    elif optimization_mode == 'max_teams':
                        result = scheduler.schedule_interviews_max_teams()
                    else:  # interviewer_friendly
                        result = scheduler.schedule_interviews_interviewer_friendly()
                    
                    # 결과 처리
                    if isinstance(result, tuple) and len(result) == 2:
                        schedule, stats = result
                    else:
                        schedule = result
                        # 기본 통계 생성
                        assigned_count = 0
                        for slot in scheduler.time_slots.values():
                            if slot.group_a_team or slot.group_b_team:
                                assigned_count += (1 if slot.group_a_team else 0)
                                assigned_count += (1 if slot.group_b_team else 0)
                        
                        stats = {
                            'assigned_teams': assigned_count,
                            'total_teams': len(st.session_state.teams),
                            'used_slots': sum(1 for slot in scheduler.time_slots.values() 
                                            if slot.group_a_team or slot.group_b_team),
                            'total_slots': len(scheduler.time_slots)
                        }
                    
                    # 결과 저장
                    st.session_state.schedule = schedule
                    st.session_state.scheduler = scheduler
                    st.session_state.stats = stats
                    
                    st.success(f"✅ 스케줄링 완료! {stats.get('assigned_teams', 0)}개 팀 배치됨")
                    
                except Exception as e:
                    st.error(f"❌ 스케줄링 중 오류 발생: {e}")
    else:
        st.warning("먼저 PDF 파일을 업로드하여 팀 정보를 등록해주세요.")

# 탭 4: A/B조 스케줄 보기
with tabs[3]:
    st.markdown("### 📋 A/B조 최종 스케줄")
    
    if st.session_state.schedule and st.session_state.scheduler:
        # 전체 통계
        stats = st.session_state.stats
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("📊 총 시간대", stats.get('total_slots', 0))
        with col2:
            st.metric("🅰️ A그룹", "배치됨")
        with col3:
            st.metric("🅱️ B그룹", "배치됨")  
        with col4:
            st.metric("✅ 총 배치", stats.get('assigned_teams', 0))
        
        st.markdown("---")
        
        # 스케줄 테이블 데이터 생성
        schedule_data = []
        
        # TimeSlot 객체에서 데이터 추출
        for slot_key, slot_obj in st.session_state.scheduler.time_slots.items():
            # slot_key에서 날짜와 시간 분리
            parts = slot_key.split(' ')
            if len(parts) >= 2:
                date = parts[0]
                time = ' '.join(parts[1:])
                
                # 배치된 팀이 있는 경우만 테이블에 추가
                a_team = getattr(slot_obj, 'group_a_team', None) or ''
                b_team = getattr(slot_obj, 'group_b_team', None) or ''
                
                if a_team or b_team:  # A조나 B조 중 하나라도 배치된 경우
                    schedule_data.append({
                        '날짜': date,
                        '시간': time,
                        'A그룹': a_team if a_team else '-',
                        'B그룹': b_team if b_team else '-'
                    })
        
        if schedule_data:
            # 날짜와 시간으로 정렬
            schedule_data.sort(key=lambda x: (x['날짜'], x['시간']))
            
            # 전체 스케줄 테이블 표시
            st.markdown("### 📊 전체 스케줄 테이블")
            df_schedule = pd.DataFrame(schedule_data)
            st.dataframe(df_schedule, use_container_width=True, hide_index=True)
            
            # 날짜별 요약 통계
            st.markdown("### 📈 날짜별 통계")
            date_names = {"9/12": "9월 12일 (금)", "9/13": "9월 13일 (토)", "9/14": "9월 14일 (일)"}
            
            summary_data = []
            for date in ["9/12", "9/13", "9/14"]:
                date_entries = [entry for entry in schedule_data if entry['날짜'] == date]
                if date_entries:
                    a_count = sum(1 for entry in date_entries if entry['A그룹'] != '-')
                    b_count = sum(1 for entry in date_entries if entry['B그룹'] != '-')
                    total_slots = len(date_entries)
                    utilization = ((a_count + b_count) / (total_slots * 2) * 100) if total_slots > 0 else 0
                    
                    summary_data.append({
                        '날짜': date_names.get(date, date),
                        '시간대수': total_slots,
                        'A그룹': f"{a_count}팀",
                        'B그룹': f"{b_count}팀",
                        '활용률': f"{utilization:.1f}%"
                    })
            
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                st.dataframe(df_summary, use_container_width=True, hide_index=True)
        else:
            st.info("배치된 스케줄이 없습니다.")
        
        # 빈 시간대 분석
        if hasattr(st.session_state.scheduler, 'calculate_gaps'):
            gaps = st.session_state.scheduler.calculate_gaps()
            if gaps:
                st.markdown("### ⚠️ 빈 시간대 분석")
                gaps_df = pd.DataFrame(gaps)
                st.dataframe(gaps_df, use_container_width=True)
    else:
        st.info("먼저 스케줄링을 실행해주세요.")

# 탭 5: 매트릭스 뷰
with tabs[4]:
    st.markdown("### 📊 팀별 가능 시간 매트릭스")
    
    if st.session_state.teams:
        # 시간대 정의
        dates = ["9/12 (금)", "9/13 (토)", "9/14 (일)"]
        time_slots_by_date = {
            "9/12 (금)": ["19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"],
            "9/13 (토)": ["10:00-10:45", "10:45-11:30", "11:30-12:15", "12:15-13:00",
                         "13:00-13:45", "13:45-14:30", "14:30-15:15", "15:15-16:00",
                         "16:00-16:45", "16:45-17:30", "17:30-18:15", "18:15-19:00",
                         "19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"],
            "9/14 (일)": ["10:00-10:45", "10:45-11:30", "11:30-12:15", "12:15-13:00",
                         "13:00-13:45", "13:45-14:30", "14:30-15:15", "15:15-16:00",
                         "16:00-16:45", "16:45-17:30", "17:30-18:15", "18:15-19:00",
                         "19:00-19:45", "19:45-20:30", "20:30-21:15", "21:15-22:00"]
        }
        
        # 날짜별 탭
        date_tabs = st.tabs(dates)
        
        for date_idx, date in enumerate(dates):
            with date_tabs[date_idx]:
                date_key = date.split(' ')[0]  # "9/12"
                time_slots = time_slots_by_date[date]
                
                # 매트릭스 데이터 생성
                matrix_data = []
                team_names = list(st.session_state.teams.keys())
                
                for team_name in team_names:
                    row = {'팀명': team_name}
                    team_info = st.session_state.teams[team_name]
                    
                    for time_slot in time_slots:
                        full_slot = f"{date_key} {time_slot}"
                        if full_slot in team_info.get('가능시간', []):
                            row[time_slot] = '⭕'
                        else:
                            row[time_slot] = ''
                    
                    matrix_data.append(row)
                
                # 데이터프레임 생성
                if matrix_data:
                    df_matrix = pd.DataFrame(matrix_data)
                    st.dataframe(df_matrix, use_container_width=True)
                    
                    # 통계
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("총 팀 수", len(team_names))
                    with col2:
                        st.metric("총 시간대", len(time_slots))
    else:
        st.info("팀 데이터가 없습니다.")

# 탭 6: 엑셀 다운로드
with tabs[5]:
    st.markdown("### 💾 엑셀 파일 다운로드")
    
    if st.session_state.schedule and st.session_state.scheduler:
        if st.button("📥 엑셀 파일 생성", key="create_excel"):
            try:
                output = BytesIO()
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "면접 스케줄"
                
                # 헤더
                headers = ["날짜", "시간", "A그룹 팀명", "B그룹 팀명"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 데이터 작성
                row_num = 2
                for slot_key, slot_obj in st.session_state.scheduler.time_slots.items():
                    if hasattr(slot_obj, 'group_a_team') or hasattr(slot_obj, 'group_b_team'):
                        parts = slot_key.split(' ')
                        if len(parts) >= 2:
                            date = parts[0]
                            time = ' '.join(parts[1:])
                            
                            ws.cell(row=row_num, column=1, value=date)
                            ws.cell(row=row_num, column=2, value=time)
                            ws.cell(row=row_num, column=3, value=getattr(slot_obj, 'group_a_team', '') or '')
                            ws.cell(row=row_num, column=4, value=getattr(slot_obj, 'group_b_team', '') or '')
                            row_num += 1
                
                # 파일 저장
                wb.save(output)
                output.seek(0)
                
                # 다운로드 버튼
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=output.getvalue(),
                    file_name=f"면접_스케줄_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ 엑셀 파일이 준비되었습니다!")
                
            except Exception as e:
                st.error(f"❌ 엑셀 파일 생성 중 오류: {e}")
    else:
        st.info("먼저 스케줄링을 실행해주세요.")

# 사이드바 상태 정보
st.sidebar.markdown("---")
st.sidebar.markdown("### 📊 현재 상태")
st.sidebar.info(f"👥 등록된 팀: {len(st.session_state.teams)}개")
if st.session_state.schedule:
    st.sidebar.success("✅ 스케줄링 완료")
    if st.session_state.stats:
        st.sidebar.info(f"📋 배치된 팀: {st.session_state.stats.get('assigned_teams', 0)}개")
else:
    st.sidebar.warning("⏳ 스케줄링 대기")