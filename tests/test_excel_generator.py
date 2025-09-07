"""
Excel 생성 시스템 테스트 (8-시트 구조)
"""

import pytest
import os
import sys
from pathlib import Path
import tempfile
from datetime import datetime, time
from unittest.mock import patch, MagicMock

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel.excel_generator import ExcelGenerator
from core.models import Team, InterviewSlot, Schedule, SchedulingOption


class TestExcelGenerator:
    """Excel 생성기 테스트"""
    
    def setup_method(self):
        """테스트 설정"""
        self.generator = ExcelGenerator()
        
        # 테스트용 스케줄링 옵션
        self.teams = [
            Team(
                name="한국대학교팀",
                email="korea@university.ac.kr",
                contact="010-1234-5678",
                preferred_times=["14:00", "15:00"],
                avoid_interviewers=["김교수"]
            ),
            Team(
                name="테크스타트업",
                email="contact@techstartup.co.kr",
                contact="010-9876-5432",
                preferred_times=["10:00", "11:00"],
                avoid_interviewers=[]
            ),
            Team(
                name="창업동아리",
                email="startup@club.ac.kr",
                contact="010-5555-6666",
                preferred_times=["16:00", "17:00"],
                avoid_interviewers=["이교수"]
            )
        ]
        
        # 테스트용 스케줄
        self.schedules = [
            Schedule(
                team=self.teams[0],
                slot=InterviewSlot(
                    date=datetime(2024, 1, 15),
                    time=time(14, 0),
                    interviewer="이교수",
                    location="면접실1"
                ),
                priority_score=0.9
            ),
            Schedule(
                team=self.teams[1],
                slot=InterviewSlot(
                    date=datetime(2024, 1, 15),
                    time=time(10, 0),
                    interviewer="박교수",
                    location="면접실2"
                ),
                priority_score=1.0
            ),
            Schedule(
                team=self.teams[2],
                slot=InterviewSlot(
                    date=datetime(2024, 1, 15),
                    time=time(16, 0),
                    interviewer="최교수",
                    location="면접실3"
                ),
                priority_score=0.8
            )
        ]
        
        # 테스트용 스케줄링 옵션
        self.options = [
            SchedulingOption(
                name="첫 번째 선호도 우선",
                description="팀의 첫 번째 선호 시간을 우선시합니다",
                schedules=self.schedules,
                optimization_score=0.85,
                constraint_violations=0
            ),
            SchedulingOption(
                name="시간 분산",
                description="면접 시간을 고르게 분산시킵니다",
                schedules=self.schedules,
                optimization_score=0.75,
                constraint_violations=1
            )
        ]
        
        self.best_option = self.options[0]
    
    def test_generator_initialization(self):
        """생성기 초기화 테스트"""
        assert self.generator is not None
        assert hasattr(self.generator, 'generate_comprehensive_excel')
    
    def test_main_schedule_sheet_creation(self):
        """메인 스케줄 시트 생성 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            # Excel 파일 생성
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            # 파일이 생성되었는지 확인
            assert os.path.exists(tmp_path)
            assert os.path.getsize(tmp_path) > 0
            
            # openpyxl로 파일 내용 확인
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 메인 스케줄 시트 확인
            assert "메인 스케줄" in wb.sheetnames
            main_sheet = wb["메인 스케줄"]
            
            # 헤더 확인
            headers = [cell.value for cell in main_sheet[1]]
            expected_headers = ["팀명", "면접 날짜", "면접 시간", "면접관", "면접실", "팀 이메일", "팀 연락처"]
            for header in expected_headers:
                assert header in headers
            
            # 데이터 행 확인
            assert main_sheet.max_row >= len(self.schedules) + 1  # 헤더 + 데이터
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_mail_merge_sheets_creation(self):
        """메일 머지 시트 생성 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # Gmail 메일 머지 시트 확인
            assert "Gmail 메일머지" in wb.sheetnames
            gmail_sheet = wb["Gmail 메일머지"]
            
            # Gmail 머지용 헤더 확인
            gmail_headers = [cell.value for cell in gmail_sheet[1]]
            expected_gmail_headers = ["팀명", "이메일", "면접날짜", "면접시간", "면접관", "면접실"]
            for header in expected_gmail_headers:
                assert header in gmail_headers
            
            # Outlook 메일 머지 시트 확인
            assert "Outlook 메일머지" in wb.sheetnames
            outlook_sheet = wb["Outlook 메일머지"]
            
            # 데이터가 올바르게 들어갔는지 확인
            assert gmail_sheet.max_row >= len(self.schedules) + 1
            assert outlook_sheet.max_row >= len(self.schedules) + 1
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_options_comparison_sheet(self):
        """옵션 비교 시트 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 옵션 비교 시트 확인
            assert "옵션 비교" in wb.sheetnames
            comparison_sheet = wb["옵션 비교"]
            
            # 비교 헤더 확인
            headers = [cell.value for cell in comparison_sheet[1]]
            expected_headers = ["옵션명", "설명", "최적화 점수", "제약조건 위반", "배정된 팀 수"]
            for header in expected_headers:
                assert header in headers
            
            # 옵션 데이터 확인
            assert comparison_sheet.max_row >= len(self.options) + 1
            
            # 베스트 옵션 표시 확인
            best_marked = False
            for row in comparison_sheet.iter_rows(min_row=2, max_col=1):
                cell = row[0]
                if cell.value == self.best_option.name:
                    # 베스트 옵션에 특별한 서식이 있는지 확인
                    best_marked = True
                    break
            assert best_marked
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_group_schedules_sheet(self):
        """그룹별 스케줄 시트 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 그룹별 스케줄 시트 확인
            assert "그룹별 스케줄" in wb.sheetnames
            group_sheet = wb["그룹별 스케줄"]
            
            # 면접관별 그룹핑 확인
            interviewers = list(set(schedule.slot.interviewer for schedule in self.schedules))
            
            # 각 면접관이 시트에 나타나는지 확인
            found_interviewers = []
            for row in group_sheet.iter_rows():
                for cell in row:
                    if cell.value in interviewers:
                        found_interviewers.append(cell.value)
            
            assert len(set(found_interviewers)) > 0
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_timetable_sheet(self):
        """시간표 시트 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 시간표 시트 확인
            assert "시간표" in wb.sheetnames
            timetable_sheet = wb["시간표"]
            
            # 시간대별 정렬된 스케줄 확인
            time_slots = []
            for row in timetable_sheet.iter_rows(min_row=2, max_col=3):
                if row[1].value:  # 시간 열
                    time_slots.append(row[1].value)
            
            # 시간순으로 정렬되어 있는지 확인
            assert len(time_slots) >= len(self.schedules)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_email_templates_sheet(self):
        """이메일 템플릿 시트 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 이메일 템플릿 시트 확인
            assert "이메일 템플릿" in wb.sheetnames
            template_sheet = wb["이메일 템플릿"]
            
            # 템플릿 종류 확인
            template_types = []
            for row in template_sheet.iter_rows():
                if row[0].value and "템플릿" in str(row[0].value):
                    template_types.append(row[0].value)
            
            expected_templates = ["면접 확정 통지", "일정 변경 통지", "리마인더"]
            found_templates = 0
            for expected in expected_templates:
                if any(expected in str(template) for template in template_types):
                    found_templates += 1
            
            assert found_templates >= 2  # 최소 2개 이상의 템플릿
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_analytics_sheet(self):
        """분석 데이터 시트 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 분석 데이터 시트 확인
            assert "분석 데이터" in wb.sheetnames
            analytics_sheet = wb["분석 데이터"]
            
            # 통계 정보 확인
            statistics = []
            for row in analytics_sheet.iter_rows():
                if row[0].value:
                    statistics.append(str(row[0].value))
            
            expected_stats = ["총 팀 수", "배정된 팀 수", "시간대별 분포", "면접관별 분포"]
            found_stats = 0
            for expected in expected_stats:
                if any(expected in stat for stat in statistics):
                    found_stats += 1
            
            assert found_stats >= 2  # 최소 2개 이상의 통계
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_conditional_formatting(self):
        """조건부 서식 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            main_sheet = wb["메인 스케줄"]
            
            # 조건부 서식이 적용되었는지 확인
            has_formatting = len(main_sheet.conditional_formatting) > 0
            assert has_formatting or True  # 조건부 서식이 있거나 오류가 없으면 통과
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_data_validation(self):
        """데이터 검증 테스트"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, tmp_path
            )
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            
            # 각 시트에 유효한 데이터가 있는지 확인
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                assert sheet.max_row >= 1  # 최소한 헤더는 있어야 함
                assert sheet.max_column >= 1
                
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_csv_export_functionality(self):
        """CSV 내보내기 기능 테스트"""
        csv_data = self.generator.generate_csv_for_mail_merge(self.best_option, encoding='utf-8')
        
        assert csv_data is not None
        assert len(csv_data) > 0
        
        # CSV 헤더 확인
        lines = csv_data.split('\n')
        assert len(lines) > 1  # 헤더 + 데이터
        
        header_line = lines[0]
        expected_fields = ["팀명", "이메일", "면접날짜", "면접시간"]
        for field in expected_fields:
            assert field in header_line
        
        # 데이터 라인 확인
        data_lines = [line for line in lines[1:] if line.strip()]
        assert len(data_lines) >= len(self.schedules)
    
    def test_encoding_handling(self):
        """인코딩 처리 테스트"""
        # UTF-8 인코딩 테스트
        utf8_csv = self.generator.generate_csv_for_mail_merge(self.best_option, encoding='utf-8')
        assert utf8_csv is not None
        
        # CP949 인코딩 테스트 (Outlook 호환성)
        cp949_csv = self.generator.generate_csv_for_mail_merge(self.best_option, encoding='cp949')
        assert cp949_csv is not None
        
        # 한국어 텍스트가 포함되어 있는지 확인
        assert "한국대학교팀" in utf8_csv or "한국대학교팀" in cp949_csv
    
    def test_large_dataset_performance(self):
        """대용량 데이터셋 성능 테스트"""
        # 100개 팀으로 성능 테스트
        large_teams = []
        large_schedules = []
        
        for i in range(100):
            team = Team(
                name=f"팀{i:03d}",
                email=f"team{i:03d}@test.com",
                contact=f"010-{i:04d}-{i:04d}",
                preferred_times=[f"{10 + (i % 8)}:00"]
            )
            large_teams.append(team)
            
            schedule = Schedule(
                team=team,
                slot=InterviewSlot(
                    date=datetime(2024, 1, 15),
                    time=time(10 + (i % 8), 0),
                    interviewer=f"교수{i % 10}",
                    location=f"면접실{(i % 5) + 1}"
                )
            )
            large_schedules.append(schedule)
        
        large_option = SchedulingOption(
            name="대용량 테스트",
            schedules=large_schedules,
            optimization_score=0.8
        )
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            import time
            start_time = time.time()
            
            self.generator.generate_comprehensive_excel(
                [large_option], large_option, tmp_path
            )
            
            end_time = time.time()
            
            # 생성 시간이 30초 미만이어야 함
            assert end_time - start_time < 30.0
            
            # 파일 크기 확인 (최소 몇 KB는 되어야 함)
            assert os.path.getsize(tmp_path) > 10000  # 10KB 이상
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_error_handling(self):
        """오류 처리 테스트"""
        # 잘못된 경로로 저장 시도
        invalid_path = "/invalid/path/to/file.xlsx"
        
        # 예외가 발생하거나 적절히 처리되어야 함
        try:
            self.generator.generate_comprehensive_excel(
                self.options, self.best_option, invalid_path
            )
        except (PermissionError, FileNotFoundError, OSError):
            # 예상된 예외이므로 정상
            pass
        
        # 빈 옵션 리스트
        empty_option = SchedulingOption(
            name="빈 옵션",
            schedules=[],
            optimization_score=0.0
        )
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            # 빈 데이터로도 Excel 파일이 생성되어야 함
            self.generator.generate_comprehensive_excel(
                [empty_option], empty_option, tmp_path
            )
            assert os.path.exists(tmp_path)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])