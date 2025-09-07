"""
전체 시스템 통합 테스트
PDF 추출 → 스케줄링 → Excel 생성 → 이메일 검증 전체 워크플로우 테스트
"""

import pytest
import sys
import os
import tempfile
from pathlib import Path
from datetime import datetime, time
from unittest.mock import patch, MagicMock

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.pdf_extractor import PDFExtractor
from core.scheduler_engine import SchedulerEngine
from excel.excel_generator import ExcelGenerator
from email.email_validator import EmailValidator
from email.template_manager import EmailTemplateManager
from core.models import Team, InterviewSlot, Schedule, SchedulingOption


class TestCompleteWorkflow:
    """완전한 워크플로우 통합 테스트"""
    
    def setup_method(self):
        """테스트 설정"""
        self.pdf_extractor = PDFExtractor()
        self.scheduler = SchedulerEngine()
        self.excel_generator = ExcelGenerator()
        self.email_validator = EmailValidator()
        self.template_manager = EmailTemplateManager()
        
        # 테스트용 면접 슬롯 생성
        self.available_slots = []
        interviewers = ["김교수", "이교수", "박교수", "최교수"]
        times = [time(9, 0), time(10, 0), time(11, 0), time(13, 0), 
                time(14, 0), time(15, 0), time(16, 0), time(17, 0)]
        
        for interviewer in interviewers:
            for slot_time in times:
                self.available_slots.append(InterviewSlot(
                    date=datetime(2024, 1, 15),
                    time=slot_time,
                    interviewer=interviewer,
                    location=f"면접실{interviewers.index(interviewer) + 1}"
                ))
    
    @patch('pdfplumber.open')
    def test_full_workflow_integration(self, mock_pdf_open):
        """전체 워크플로우 통합 테스트"""
        # 1단계: PDF 데이터 모킹
        mock_page = MagicMock()
        sample_pdf_content = """
        팀명: 한국대학교 AI팀
        이메일: ai.team@korea.ac.kr
        연락처: 010-1111-2222
        선호시간: 14:00, 15:00
        피하고싶은 면접관: 김교수
        
        팀명: 스타트업 혁신팀
        이메일: innovation@startup.co.kr
        연락처: 010-3333-4444
        선호시간: 10:00, 11:00
        피하고싶은 면접관: 없음
        
        팀명: 테크 솔루션팀
        이메일: tech.solution@company.com
        연락처: 010-5555-6666
        선호시간: 16:00, 17:00
        피하고싶은 면접관: 이교수
        """
        
        mock_page.extract_text.return_value = sample_pdf_content
        mock_page.extract_tables.return_value = []
        
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__.return_value = mock_pdf
        mock_pdf_open.return_value = mock_pdf
        
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_pdf:
            pdf_path = tmp_pdf.name
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_excel:
            excel_path = tmp_excel.name
        
        try:
            # 2단계: PDF에서 팀 데이터 추출
            teams = self.pdf_extractor.extract_team_data(pdf_path)
            
            assert len(teams) >= 3, "Should extract at least 3 teams"
            team_names = [team.name for team in teams]
            assert "한국대학교 AI팀" in team_names
            assert "스타트업 혁신팀" in team_names
            assert "테크 솔루션팀" in team_names
            
            # 3단계: 이메일 검증
            validated_teams = []
            for team in teams:
                validation_result = self.email_validator.validate_email(team.email)
                if validation_result["is_valid"]:
                    validated_teams.append(team)
                else:
                    print(f"Invalid email for team {team.name}: {team.email}")
            
            assert len(validated_teams) >= 2, "At least 2 teams should have valid emails"
            
            # 4단계: 스케줄링 최적화
            scheduling_options = self.scheduler.generate_five_options(
                validated_teams, self.available_slots
            )
            
            assert len(scheduling_options) == 5, "Should generate 5 scheduling options"
            best_option = max(scheduling_options, key=lambda x: x.optimization_score)
            assert best_option.optimization_score > 0, "Best option should have positive score"
            
            # 5단계: Excel 파일 생성
            self.excel_generator.generate_comprehensive_excel(
                scheduling_options, best_option, excel_path
            )
            
            assert os.path.exists(excel_path), "Excel file should be created"
            assert os.path.getsize(excel_path) > 0, "Excel file should not be empty"
            
            # 6단계: Excel 파일 검증
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            
            expected_sheets = [
                "메인 스케줄", "Gmail 메일머지", "Outlook 메일머지", 
                "옵션 비교", "그룹별 스케줄", "시간표", "이메일 템플릿", "분석 데이터"
            ]
            
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"
            
            # 메인 스케줄 시트 데이터 확인
            main_sheet = wb["메인 스케줄"]
            assert main_sheet.max_row > 1, "Main schedule should have data rows"
            
            # 7단계: 이메일 템플릿 생성
            for schedule in best_option.schedules[:3]:  # 처음 3개만 테스트
                email_content = self.template_manager.generate_email(
                    template_type="interview_notification",
                    schedule=schedule
                )
                
                assert email_content["subject"], "Email should have subject"
                assert email_content["body"], "Email should have body"
                assert email_content["recipient"] == schedule.team.email
                assert schedule.team.name in email_content["body"]
            
            # 8단계: 제약조건 검증
            for schedule in best_option.schedules:
                assert schedule.is_valid(), f"Schedule should be valid for {schedule.team.name}"
                
                # 팀이 피하는 면접관이 배정되지 않았는지 확인
                if schedule.team.avoid_interviewers:
                    assert schedule.slot.interviewer not in schedule.team.avoid_interviewers, \
                        f"Team {schedule.team.name} assigned to avoided interviewer {schedule.slot.interviewer}"
            
            print(f"✅ 전체 워크플로우 테스트 성공!")
            print(f"   - 추출된 팀: {len(teams)}개")
            print(f"   - 검증된 팀: {len(validated_teams)}개") 
            print(f"   - 생성된 옵션: {len(scheduling_options)}개")
            print(f"   - 최적화 점수: {best_option.optimization_score:.3f}")
            print(f"   - 제약조건 위반: {best_option.constraint_violations}개")
            
        finally:
            # 임시 파일 정리
            for path in [pdf_path, excel_path]:
                if os.path.exists(path):
                    os.unlink(path)
    
    def test_error_recovery_workflow(self):
        """오류 복구 워크플로우 테스트"""
        # 잘못된 PDF 경로
        invalid_pdf_path = "/non/existent/file.pdf"
        teams = self.pdf_extractor.extract_team_data(invalid_pdf_path)
        
        # 빈 팀 리스트로도 스케줄링이 작동해야 함
        assert teams == [], "Should return empty list for invalid PDF"
        
        # 빈 팀으로 스케줄링 시도
        empty_options = self.scheduler.generate_five_options(teams, self.available_slots)
        assert len(empty_options) == 5, "Should still generate 5 options even with empty teams"
        
        # 빈 옵션으로 Excel 생성
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            excel_path = tmp_file.name
        
        try:
            self.excel_generator.generate_comprehensive_excel(
                empty_options, empty_options[0], excel_path
            )
            assert os.path.exists(excel_path), "Should create Excel even with empty data"
            
        finally:
            if os.path.exists(excel_path):
                os.unlink(excel_path)
    
    def test_large_scale_workflow(self):
        """대규모 워크플로우 테스트"""
        # 70개 팀 데이터 생성 (실제 요구사항)
        large_teams = []
        for i in range(70):
            team = Team(
                name=f"팀{i+1:02d}호",
                email=f"team{i+1:02d}@test{(i % 10) + 1}.com",
                contact=f"010-{i+1:04d}-{i+1:04d}",
                preferred_times=[f"{9 + (i % 9)}:00", f"{10 + (i % 9)}:00"],
                avoid_interviewers=[f"교수{(i % 4) + 1}"] if i % 5 == 0 else []
            )
            large_teams.append(team)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            excel_path = tmp_file.name
        
        try:
            import time
            start_time = time.time()
            
            # 대규모 스케줄링
            large_options = self.scheduler.generate_five_options(large_teams, self.available_slots)
            
            # 스케줄링 시간 측정
            scheduling_time = time.time() - start_time
            
            assert len(large_options) == 5
            assert scheduling_time < 120.0, f"Large scale scheduling took too long: {scheduling_time:.2f}s"
            
            # 최적 옵션 선택
            best_large_option = max(large_options, key=lambda x: x.optimization_score)
            
            # Excel 생성
            excel_start_time = time.time()
            self.excel_generator.generate_comprehensive_excel(
                large_options, best_large_option, excel_path
            )
            excel_time = time.time() - excel_start_time
            
            assert os.path.exists(excel_path)
            assert excel_time < 60.0, f"Excel generation took too long: {excel_time:.2f}s"
            
            # 파일 크기 확인 (대규모 데이터)
            file_size = os.path.getsize(excel_path)
            assert file_size > 50000, f"Excel file too small for large dataset: {file_size} bytes"
            
            print(f"✅ 대규모 워크플로우 테스트 성공!")
            print(f"   - 팀 수: {len(large_teams)}개")
            print(f"   - 스케줄링 시간: {scheduling_time:.2f}초")
            print(f"   - Excel 생성 시간: {excel_time:.2f}초")
            print(f"   - 배정된 팀: {len(best_large_option.schedules)}개")
            print(f"   - 파일 크기: {file_size:,} bytes")
            
        finally:
            if os.path.exists(excel_path):
                os.unlink(excel_path)
    
    def test_constraint_satisfaction_integration(self):
        """제약조건 만족도 통합 테스트"""
        # 복잡한 제약조건을 가진 팀들
        constraint_teams = [
            Team(
                name="까다로운팀1",
                email="difficult1@test.com",
                contact="010-1111-1111",
                preferred_times=["14:00"],  # 특정 시간만 선호
                avoid_interviewers=["김교수", "이교수"]  # 여러 면접관 회피
            ),
            Team(
                name="까다로운팀2",
                email="difficult2@test.com",
                contact="010-2222-2222",
                preferred_times=["14:00"],  # 동일 시간 선호 (충돌)
                avoid_interviewers=["박교수"]
            ),
            Team(
                name="유연한팀",
                email="flexible@test.com",
                contact="010-3333-3333",
                preferred_times=["09:00", "10:00", "11:00", "13:00", "15:00"],
                avoid_interviewers=[]
            )
        ]
        
        # 제한된 슬롯 (14:00에 김교수, 이교수만 있다고 가정)
        limited_slots = [
            InterviewSlot(
                date=datetime(2024, 1, 15),
                time=time(14, 0),
                interviewer="김교수",
                location="면접실1"
            ),
            InterviewSlot(
                date=datetime(2024, 1, 15),
                time=time(14, 0),
                interviewer="이교수",
                location="면접실2"
            ),
            InterviewSlot(
                date=datetime(2024, 1, 15),
                time=time(14, 0),
                interviewer="박교수",
                location="면접실3"
            ),
            InterviewSlot(
                date=datetime(2024, 1, 15),
                time=time(15, 0),
                interviewer="최교수",
                location="면접실4"
            )
        ]
        
        constraint_options = self.scheduler.generate_five_options(constraint_teams, limited_slots)
        
        # 제약조건 만족도 검증
        for option in constraint_options:
            for schedule in option.schedules:
                # 면접관 회피 제약조건 확인
                assert schedule.slot.interviewer not in schedule.team.avoid_interviewers, \
                    f"Constraint violation: {schedule.team.name} assigned to {schedule.slot.interviewer}"
        
        # 최적 솔루션에서 가능한 많은 팀이 배정되었는지 확인
        best_constraint_option = max(constraint_options, key=lambda x: len(x.schedules))
        
        # 유연한팀은 반드시 배정되어야 함
        flexible_team_assigned = any(
            schedule.team.name == "유연한팀" 
            for schedule in best_constraint_option.schedules
        )
        assert flexible_team_assigned, "Flexible team should always be assigned"
        
        print(f"✅ 제약조건 만족도 테스트 성공!")
        print(f"   - 까다로운 팀들 중 배정된 수: {len(best_constraint_option.schedules)}개")
    
    def test_performance_benchmarking(self):
        """성능 벤치마킹 테스트"""
        import time
        import psutil
        import os
        
        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss
        
        # 중간 규모 데이터 (30개 팀)
        medium_teams = []
        for i in range(30):
            team = Team(
                name=f"벤치마크팀{i+1}",
                email=f"benchmark{i+1}@test.com",
                contact=f"010-{i:04d}-{i:04d}",
                preferred_times=[f"{10 + (i % 6)}:00"],
                avoid_interviewers=[f"교수{i % 3}"] if i % 4 == 0 else []
            )
            medium_teams.append(team)
        
        # 성능 측정
        start_time = time.time()
        
        # 이메일 검증 (병렬)
        email_start = time.time()
        email_results = self.email_validator.validate_bulk(
            [team.email for team in medium_teams], 
            max_workers=4
        )
        email_time = time.time() - email_start
        
        # 스케줄링 최적화
        scheduling_start = time.time()
        benchmark_options = self.scheduler.generate_five_options(medium_teams, self.available_slots)
        scheduling_time = time.time() - scheduling_start
        
        # Excel 생성
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            excel_path = tmp_file.name
        
        try:
            excel_start = time.time()
            self.excel_generator.generate_comprehensive_excel(
                benchmark_options, benchmark_options[0], excel_path
            )
            excel_time = time.time() - excel_start
            
            total_time = time.time() - start_time
            final_memory = process.memory_info().rss
            memory_usage = final_memory - initial_memory
            
            # 성능 기준 검증
            assert email_time < 10.0, f"Email validation too slow: {email_time:.2f}s"
            assert scheduling_time < 60.0, f"Scheduling too slow: {scheduling_time:.2f}s"
            assert excel_time < 30.0, f"Excel generation too slow: {excel_time:.2f}s"
            assert total_time < 90.0, f"Total workflow too slow: {total_time:.2f}s"
            assert memory_usage < 200 * 1024 * 1024, f"Memory usage too high: {memory_usage/1024/1024:.1f}MB"
            
            print(f"✅ 성능 벤치마킹 테스트 성공!")
            print(f"   - 이메일 검증: {email_time:.2f}초")
            print(f"   - 스케줄링: {scheduling_time:.2f}초") 
            print(f"   - Excel 생성: {excel_time:.2f}초")
            print(f"   - 총 시간: {total_time:.2f}초")
            print(f"   - 메모리 사용량: {memory_usage/1024/1024:.1f}MB")
            
        finally:
            if os.path.exists(excel_path):
                os.unlink(excel_path)
    
    def test_data_consistency_across_modules(self):
        """모듈 간 데이터 일관성 테스트"""
        # 샘플 팀 데이터
        sample_teams = [
            Team(
                name="일관성테스트팀",
                email="consistency@test.com",
                contact="010-9999-9999",
                preferred_times=["14:00"],
                avoid_interviewers=[]
            )
        ]
        
        # 스케줄링
        consistency_options = self.scheduler.generate_five_options(sample_teams, self.available_slots)
        best_consistency_option = consistency_options[0]
        
        # Excel 생성 및 검증
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            excel_path = tmp_file.name
        
        try:
            self.excel_generator.generate_comprehensive_excel(
                consistency_options, best_consistency_option, excel_path
            )
            
            # Excel에서 데이터 읽기
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            main_sheet = wb["메인 스케줄"]
            
            # 첫 번째 데이터 행 확인
            if main_sheet.max_row > 1:
                first_row = list(main_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                excel_team_name = first_row[0]  # 첫 번째 컬럼: 팀명
                excel_email = first_row[5]      # 여섯 번째 컬럼: 이메일
                
                # 원본 데이터와 일치 확인
                original_team = sample_teams[0]
                assert excel_team_name == original_team.name, "Team name mismatch in Excel"
                assert excel_email == original_team.email, "Email mismatch in Excel"
            
            # CSV 생성 및 일관성 확인
            csv_data = self.excel_generator.generate_csv_for_mail_merge(
                best_consistency_option, encoding='utf-8'
            )
            
            assert "일관성테스트팀" in csv_data, "Team name should be in CSV"
            assert "consistency@test.com" in csv_data, "Email should be in CSV"
            
            # 이메일 템플릿 생성 및 일관성 확인
            if best_consistency_option.schedules:
                schedule = best_consistency_option.schedules[0]
                email_content = self.template_manager.generate_email(
                    template_type="interview_notification",
                    schedule=schedule
                )
                
                assert schedule.team.name in email_content["body"], "Team name should be in email"
                assert email_content["recipient"] == schedule.team.email, "Email recipient mismatch"
            
            print(f"✅ 데이터 일관성 테스트 성공!")
            
        finally:
            if os.path.exists(excel_path):
                os.unlink(excel_path)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])  # -s 옵션으로 print 출력 표시