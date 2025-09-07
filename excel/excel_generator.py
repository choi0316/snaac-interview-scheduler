"""
엑셀 생성 시스템 - 8개 시트 구조 및 메일 머지 최적화

공모전 면접 스케줄링 결과를 종합적인 엑셀 파일로 출력하는 모듈입니다.
메일 발송에 최적화된 구조와 다양한 분석 시트를 포함합니다.
"""

import logging
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
import io

from ..core.models import Schedule, SchedulingOption, SchedulingResult

logger = logging.getLogger(__name__)


class ExcelStyleManager:
    """엑셀 스타일 관리자"""
    
    def __init__(self):
        self.styles = {}
        self._create_styles()
    
    def _create_styles(self):
        """기본 스타일 생성"""
        
        # 헤더 스타일
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
        header_style.fill = PatternFill("solid", fgColor="366092")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.styles['header'] = header_style
        
        # 데이터 스타일
        data_style = NamedStyle(name="data")
        data_style.font = Font(name='맑은 고딕', size=10)
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.styles['data'] = data_style
        
        # 숫자 스타일
        number_style = NamedStyle(name="number")
        number_style.font = Font(name='맑은 고딕', size=10)
        number_style.alignment = Alignment(horizontal="right", vertical="center")
        number_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.styles['number'] = number_style
        
        # 날짜 스타일
        date_style = NamedStyle(name="date")
        date_style.font = Font(name='맑은 고딕', size=10)
        date_style.alignment = Alignment(horizontal="center", vertical="center")
        date_style.number_format = "YYYY-MM-DD"
        date_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.styles['date'] = date_style


class ExcelGenerator:
    """종합 엑셀 파일 생성기"""
    
    def __init__(self):
        self.wb = None
        self.style_manager = ExcelStyleManager()
        
        # 메일 머지 컬럼 정의
        self.mail_merge_columns = [
            ('team_name', '팀명'),
            ('leader_name', '대표자명'),
            ('email', '이메일'),
            ('phone', '연락처'),
            ('interview_date', '면접날짜'),
            ('interview_time', '면접시간'),
            ('interview_group', '면접조'),
            ('interview_room', '면접장소'),
            ('zoom_link', '줌링크'),
            ('status', '합격여부'),
            ('mail_sent', '메일발송상태'),
            ('notes', '비고')
        ]
        
        # 메일 템플릿 변수
        self.template_variables = {
            '{{team_name}}': '팀명',
            '{{leader_name}}': '대표자명',
            '{{interview_date}}': '면접날짜', 
            '{{interview_time}}': '면접시간',
            '{{interview_room}}': '면접장소',
            '{{zoom_link}}': '줌링크',
            '{{additional_info}}': '추가안내사항'
        }
    
    def generate_complete_excel(
        self,
        scheduling_result: SchedulingResult,
        output_path: Optional[str] = None
    ) -> Tuple[Workbook, str]:
        """완전한 엑셀 파일 생성 (8개 시트)"""
        
        logger.info("8개 시트 엑셀 파일 생성 시작")
        
        # 새 워크북 생성
        self.wb = Workbook()
        
        # 기본 시트 제거
        if self.wb.worksheets:
            self.wb.remove(self.wb.active)
        
        # 스타일 등록
        for style in self.style_manager.styles.values():
            if style.name not in [s.name for s in self.wb.named_styles]:
                self.wb.add_named_style(style)
        
        try:
            # 선택된 옵션이 없으면 첫 번째 옵션 사용
            selected_option = scheduling_result.selected_option
            if not selected_option and scheduling_result.options:
                selected_option = scheduling_result.options[0]
            
            if not selected_option:
                raise ValueError("선택된 스케줄링 옵션이 없습니다.")
            
            # 8개 시트 생성
            self._create_main_schedule_sheet(selected_option)
            self._create_mail_merge_sheets(selected_option)
            self._create_options_comparison_sheet(scheduling_result.options)
            self._create_group_schedule_sheets(selected_option)
            self._create_timetable_sheet(selected_option)
            self._create_email_template_sheet()
            self._create_analytics_sheet(selected_option)
            
            # 파일 저장
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"interview_schedule_{timestamp}.xlsx"
            
            self.wb.save(output_path)
            logger.info(f"엑셀 파일 생성 완료: {output_path}")
            
            return self.wb, output_path
            
        except Exception as e:
            logger.error(f"엑셀 파일 생성 오류: {e}")
            raise
    
    def _create_main_schedule_sheet(self, option: SchedulingOption):
        """메인 스케줄 시트 생성 (시트 1)"""
        ws = self.wb.create_sheet("최종_스케줄")
        
        # 헤더 생성
        headers = [col[1] for col in self.mail_merge_columns]
        ws.append(headers)
        
        # 헤더 스타일 적용
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
            
        # 데이터 입력
        for schedule in option.schedules:
            row_data = self._schedule_to_row_data(schedule)
            ws.append(row_data)
        
        # 데이터 스타일 적용
        for row_num in range(2, len(option.schedules) + 2):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col_num)
                if col_num in [5, 6]:  # 날짜, 시간 컬럼
                    cell.style = 'date'
                elif col_num in [11]:  # 메일발송상태
                    cell.style = 'data'
                else:
                    cell.style = 'data'
        
        # 조건부 서식 적용
        self._apply_conditional_formatting(ws)
        
        # 열 너비 자동 조정
        self._adjust_column_widths(ws)
        
        # 필터 추가
        ws.auto_filter.ref = f"A1:{chr(ord('A') + len(headers) - 1)}{len(option.schedules) + 1}"
        
        # 데이터 검증 추가
        self._add_data_validation(ws)
    
    def _create_mail_merge_sheets(self, option: SchedulingOption):
        """메일 머지용 데이터 시트 생성 (시트 2-3)"""
        
        # 합격자용 시트
        self._create_single_mail_merge_sheet(option, "accepted", "메일머지_합격자")
        
        # 탈락자용 시트 (현재는 모두 합격으로 가정)
        self._create_single_mail_merge_sheet(option, "rejected", "메일머지_탈락자")
    
    def _create_single_mail_merge_sheet(self, option: SchedulingOption, status: str, sheet_name: str):
        """단일 메일 머지 시트 생성"""
        ws = self.wb.create_sheet(sheet_name)
        
        # 메일 머지용 헤더 (영문 필드명 + 한글 설명)
        headers = []
        for eng_name, kor_name in self.mail_merge_columns:
            headers.append(f"{kor_name}")  # Gmail/Outlook 호환을 위해 한글 헤더 사용
        
        ws.append(headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
        
        # 데이터 필터링 및 입력
        filtered_schedules = option.schedules  # 현재는 모든 팀이 합격
        if status == "rejected":
            filtered_schedules = []  # 탈락자는 별도 처리 필요
        
        for schedule in filtered_schedules:
            row_data = self._schedule_to_mail_merge_data(schedule)
            ws.append(row_data)
        
        # 스타일 및 서식 적용
        self._apply_mail_merge_formatting(ws, len(filtered_schedules))
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_options_comparison_sheet(self, options: List[SchedulingOption]):
        """5개 옵션 비교 시트 생성 (시트 4)"""
        ws = self.wb.create_sheet("옵션_비교")
        
        # 비교 테이블 헤더
        comparison_headers = [
            '옵션명', '전략유형', '총팀수', '1순위만족률(%)', 
            '조별균형점수', '제약위반수', '생성시간(초)', '추천여부'
        ]
        
        ws.append(comparison_headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(comparison_headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
        
        # 옵션별 데이터 입력
        for option in options:
            option.calculate_satisfaction_metrics()  # 지표 계산
            
            row_data = [
                option.option_name,
                option.strategy_type,
                option.total_teams,
                f"{option.first_choice_satisfaction:.1f}",
                f"{option.group_balance_score:.1f}",
                option.constraint_violations,
                f"{option.generation_time:.2f}",
                "추천" if option.first_choice_satisfaction > 80 else ""
            ]
            ws.append(row_data)
        
        # 스타일 적용
        for row_num in range(2, len(options) + 2):
            for col_num in range(1, len(comparison_headers) + 1):
                cell = ws.cell(row_num, col_num)
                if col_num in [3, 6]:  # 숫자 컬럼
                    cell.style = 'number'
                else:
                    cell.style = 'data'
        
        # 차트 생성
        self._create_comparison_chart(ws, len(options))
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _create_group_schedule_sheets(self, option: SchedulingOption):
        """A조/B조 일정표 시트 생성 (시트 5-6)"""
        
        for group_name in ["A조", "B조"]:
            ws = self.wb.create_sheet(f"{group_name}_일정")
            
            # 해당 조 스케줄만 필터링
            group_schedules = [
                s for s in option.schedules 
                if s.interview_group == group_name
            ]
            
            # 시간순 정렬
            group_schedules.sort(
                key=lambda x: (x.interview_date, x.interview_time)
            )
            
            # 테이블 생성
            headers = ['면접시간', '팀명', '대표자', '이메일', '면접실', '비고']
            ws.append(headers)
            
            # 헤더 스타일
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(1, col_num)
                cell.style = 'header'
            
            # 데이터 입력
            for schedule in group_schedules:
                row_data = [
                    f"{schedule.interview_date} {schedule.interview_time}",
                    schedule.team_name,
                    schedule.leader_name,
                    schedule.team.primary_email,
                    schedule.interview_slot.room,
                    schedule.notes or ""
                ]
                ws.append(row_data)
            
            # 스타일 적용
            for row_num in range(2, len(group_schedules) + 2):
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row_num, col_num)
                    cell.style = 'data'
            
            # 열 너비 조정
            self._adjust_column_widths(ws)
            
            # 그룹별 통계 추가
            self._add_group_statistics(ws, group_schedules, group_name)
    
    def _create_timetable_sheet(self, option: SchedulingOption):
        """시간대별 타임테이블 시트 생성 (시트 7)"""
        ws = self.wb.create_sheet("타임테이블")
        
        # 시간대별 데이터 구조화
        timetable_data = self._build_timetable_data(option.schedules)
        
        # 타임테이블 헤더 생성
        dates = sorted(set(s.interview_date for s in option.schedules))
        time_slots = sorted(set(s.interview_time for s in option.schedules))
        
        # 헤더 구성: 시간 | 날짜1-A조 | 날짜1-B조 | 날짜2-A조 | 날짜2-B조 ...
        headers = ['시간대']
        for date in dates:
            headers.extend([f"{date}-A조", f"{date}-B조"])
        
        ws.append(headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
        
        # 시간대별 데이터 입력
        for time_slot in time_slots:
            row_data = [time_slot]
            
            for date in dates:
                for group in ["A조", "B조"]:
                    team_name = timetable_data.get((date, time_slot, group), "")
                    row_data.append(team_name)
            
            ws.append(row_data)
        
        # 스타일 적용
        for row_num in range(2, len(time_slots) + 2):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col_num)
                cell.style = 'data'
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
        
        # 타임테이블 차트 생성
        self._create_timetable_chart(ws, len(time_slots), len(headers))
    
    def _create_email_template_sheet(self):
        """메일 템플릿 시트 생성 (시트 8)"""
        ws = self.wb.create_sheet("메일_템플릿")
        
        # 템플릿 데이터
        templates = [
            {
                'type': '합격자_면접안내',
                'subject': '[공모전] {{team_name}} 팀 2차 면접 안내',
                'body': """안녕하세요, {{leader_name}}님.

{{team_name}} 팀의 2차 면접 일정을 안내드립니다.

📅 면접 일정 정보
• 면접 날짜: {{interview_date}}
• 면접 시간: {{interview_time}} 
• 면접 조: {{interview_group}}
• 면접 장소: {{interview_room}}
• 온라인 링크: {{zoom_link}}

📋 면접 준비사항
• 발표자료: 10분 이내 프레젠테이션
• 질의응답: 5분
• 지참물: 신분증, 팀원 전원 참석

{{additional_info}}

문의사항이 있으시면 언제든 연락 주세요.

감사합니다."""
            },
            {
                'type': '탈락자_안내',
                'subject': '[공모전] 심사 결과 안내',
                'body': """안녕하세요, {{leader_name}}님.

{{team_name}} 팀의 공모전 참가에 감사드립니다.

아쉽게도 2차 면접 대상에서 제외되었음을 안내드립니다.
많은 팀이 지원하여 경쟁이 치열했습니다.

앞으로도 좋은 기회로 만나뵐 수 있기를 바랍니다.

감사합니다."""
            },
            {
                'type': '면접_리마인더',
                'subject': '[공모전] 내일 면접 일정 리마인더',
                'body': """안녕하세요, {{leader_name}}님.

내일 예정된 {{team_name}} 팀의 면접 일정을 리마인드드립니다.

📅 면접 일정
• 면접 시간: {{interview_time}}
• 면접 장소: {{interview_room}}

준비물과 시간을 다시 한번 확인해주세요.

감사합니다."""
            }
        ]
        
        # 테이블 헤더
        headers = ['템플릿유형', '제목', '내용', '사용변수']
        ws.append(headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
        
        # 템플릿 데이터 입력
        for template in templates:
            variables = ', '.join(self.template_variables.keys())
            row_data = [
                template['type'],
                template['subject'],
                template['body'],
                variables
            ]
            ws.append(row_data)
        
        # 스타일 적용 및 행 높이 조정
        for row_num in range(2, len(templates) + 2):
            ws.row_dimensions[row_num].height = 120  # 높이 증가
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col_num)
                cell.style = 'data'
                cell.alignment = Alignment(
                    horizontal="left", 
                    vertical="top", 
                    wrap_text=True
                )
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 30
        
        # 변수 설명 추가
        self._add_template_variable_guide(ws, len(templates) + 3)
    
    def _create_analytics_sheet(self, option: SchedulingOption):
        """분석 시트 생성 (추가)"""
        ws = self.wb.create_sheet("분석_대시보드")
        
        # 기본 통계
        stats = self._calculate_schedule_statistics(option)
        
        # 통계 테이블 생성
        stats_headers = ['지표', '값', '목표', '달성여부']
        ws.append(stats_headers)
        
        # 헤더 스타일
        for col_num, header in enumerate(stats_headers, 1):
            cell = ws.cell(1, col_num)
            cell.style = 'header'
        
        # 통계 데이터
        stats_data = [
            ['총 팀 수', stats['total_teams'], '-', '✓'],
            ['1순위 만족률', f"{stats['first_choice_rate']:.1f}%", '80%', 
             '✓' if stats['first_choice_rate'] >= 80 else '✗'],
            ['A조 팀 수', stats['group_a_count'], '50%±5%', 
             '✓' if abs(stats['group_a_count'] - stats['total_teams']/2) <= stats['total_teams']*0.05 else '✗'],
            ['B조 팀 수', stats['group_b_count'], '50%±5%', 
             '✓' if abs(stats['group_b_count'] - stats['total_teams']/2) <= stats['total_teams']*0.05 else '✗'],
            ['평균 팀원 수', f"{stats['avg_team_size']:.1f}명", '-', '✓'],
            ['제약 위반 수', stats['constraint_violations'], '0건', 
             '✓' if stats['constraint_violations'] == 0 else '✗']
        ]
        
        for row_data in stats_data:
            ws.append(row_data)
        
        # 스타일 적용
        for row_num in range(2, len(stats_data) + 2):
            for col_num in range(1, len(stats_headers) + 1):
                cell = ws.cell(row_num, col_num)
                cell.style = 'data'
        
        # 열 너비 조정
        self._adjust_column_widths(ws)
    
    def _schedule_to_row_data(self, schedule: Schedule) -> List[Any]:
        """스케줄을 행 데이터로 변환"""
        return [
            schedule.team_name,
            schedule.leader_name,
            schedule.team.primary_email,
            schedule.team.primary_phone,
            schedule.interview_date,
            schedule.interview_time,
            schedule.interview_group,
            schedule.interview_slot.room,
            schedule.interview_slot.zoom_link or "",
            "합격",  # 기본값
            "미발송",  # 기본값
            schedule.notes or ""
        ]
    
    def _schedule_to_mail_merge_data(self, schedule: Schedule) -> List[Any]:
        """스케줄을 메일 머지 데이터로 변환"""
        return self._schedule_to_row_data(schedule)
    
    def _apply_conditional_formatting(self, ws):
        """조건부 서식 적용"""
        
        # 면접 시간 임박 (빨강) - E열(면접날짜)
        red_fill = PatternFill("solid", fgColor="FF6B6B")
        urgent_rule = CellIsRule(
            operator='lessThan',
            formula=['TODAY()+1'],
            fill=red_fill
        )
        ws.conditional_formatting.add('E2:E1000', urgent_rule)
        
        # 메일 발송 완료 (초록) - K열(메일발송상태)
        green_fill = PatternFill("solid", fgColor="51CF66")
        sent_rule = CellIsRule(
            operator='equal',
            formula=['"완료"'],
            fill=green_fill
        )
        ws.conditional_formatting.add('K2:K1000', sent_rule)
        
        # 제약 조건 충돌 (노랑) - L열(비고)
        yellow_fill = PatternFill("solid", fgColor="FFE066")
        conflict_rule = CellIsRule(
            operator='containsText',
            formula=['"충돌"'],
            fill=yellow_fill
        )
        ws.conditional_formatting.add('L2:L1000', conflict_rule)
    
    def _apply_mail_merge_formatting(self, ws, data_count: int):
        """메일 머지 시트 서식 적용"""
        
        # 데이터 스타일 적용
        for row_num in range(2, data_count + 2):
            for col_num in range(1, len(self.mail_merge_columns) + 1):
                cell = ws.cell(row_num, col_num)
                cell.style = 'data'
        
        # 이메일 열 검증 추가 (C열)
        email_validation = DataValidation(
            type="custom",
            formula1='=AND(ISERROR(FIND(" ",C2))=TRUE,LEN(C2)-LEN(SUBSTITUTE(C2,"@",""))=1)',
            showErrorMessage=True,
            errorTitle="이메일 형식 오류",
            error="올바른 이메일 형식이 아닙니다."
        )
        ws.add_data_validation(email_validation)
        email_validation.add(f'C2:C{data_count + 1}')
    
    def _add_data_validation(self, ws):
        """데이터 검증 추가"""
        
        # 메일 발송 상태 검증 (K열)
        status_validation = DataValidation(
            type="list",
            formula1='"미발송,완료,실패"',
            showDropDown=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add('K2:K1000')
        
        # 합격 여부 검증 (J열)
        result_validation = DataValidation(
            type="list", 
            formula1='"합격,탈락,대기"',
            showDropDown=True
        )
        ws.add_data_validation(result_validation)
        result_validation.add('J2:J1000')
    
    def _adjust_column_widths(self, ws):
        """열 너비 자동 조정"""
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            adjusted_width = min(max(length + 2, 10), 50)  # 10-50 범위
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    def _create_comparison_chart(self, ws, option_count: int):
        """옵션 비교 차트 생성"""
        
        # 1순위 만족률 차트
        chart = BarChart()
        chart.type = "col"
        chart.title = "옵션별 1순위 만족률 비교"
        chart.y_axis.title = "만족률 (%)"
        chart.x_axis.title = "옵션"
        
        # 데이터 범위 설정
        data = Reference(ws, min_col=4, min_row=1, max_row=option_count + 1, max_col=4)
        categories = Reference(ws, min_col=1, min_row=2, max_row=option_count + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # 차트 위치 설정
        ws.add_chart(chart, f"J2")
    
    def _create_timetable_chart(self, ws, time_slot_count: int, header_count: int):
        """타임테이블 차트 생성"""
        
        # 시간대별 배치 현황 차트
        chart = BarChart()
        chart.type = "col"
        chart.title = "시간대별 면접 배치 현황"
        chart.y_axis.title = "배치 팀 수"
        chart.x_axis.title = "시간대"
        
        # 간단한 집계를 위해 차트는 생략하고 나중에 구현
        pass
    
    def _build_timetable_data(self, schedules: List[Schedule]) -> Dict[Tuple[str, str, str], str]:
        """타임테이블 데이터 구조화"""
        timetable = {}
        
        for schedule in schedules:
            key = (schedule.interview_date, schedule.interview_time, schedule.interview_group)
            timetable[key] = schedule.team_name
        
        return timetable
    
    def _add_group_statistics(self, ws, schedules: List[Schedule], group_name: str):
        """그룹별 통계 추가"""
        
        # 빈 행 추가
        ws.append([])
        ws.append([f"{group_name} 통계"])
        
        stats_row = len(schedules) + 3
        
        # 통계 정보
        total_teams = len(schedules)
        time_distribution = {}
        for schedule in schedules:
            time_key = schedule.interview_time
            time_distribution[time_key] = time_distribution.get(time_key, 0) + 1
        
        # 통계 데이터 추가
        ws.append([f"총 팀 수: {total_teams}"])
        ws.append([f"시간대 수: {len(time_distribution)}"])
        
        # 가장 많은 팀이 배치된 시간
        if time_distribution:
            max_time = max(time_distribution.items(), key=lambda x: x[1])
            ws.append([f"최대 배치 시간: {max_time[0]} ({max_time[1]}팀)"])
    
    def _add_template_variable_guide(self, ws, start_row: int):
        """템플릿 변수 가이드 추가"""
        
        ws.cell(start_row, 1, "템플릿 변수 설명")
        ws.cell(start_row, 1).style = 'header'
        
        start_row += 2
        
        for variable, description in self.template_variables.items():
            ws.append([variable, description])
    
    def _calculate_schedule_statistics(self, option: SchedulingOption) -> Dict[str, Any]:
        """스케줄 통계 계산"""
        
        if not option.schedules:
            return {}
        
        total_teams = len(option.schedules)
        group_a_count = len([s for s in option.schedules if s.interview_group == "A조"])
        group_b_count = total_teams - group_a_count
        
        # 1순위 만족률
        first_choice_count = len([s for s in option.schedules if s.preference_rank == 1])
        first_choice_rate = (first_choice_count / total_teams * 100) if total_teams > 0 else 0
        
        # 평균 팀원 수
        total_members = sum(len(s.team.members) for s in option.schedules)
        avg_team_size = total_members / total_teams if total_teams > 0 else 0
        
        return {
            'total_teams': total_teams,
            'group_a_count': group_a_count,
            'group_b_count': group_b_count,
            'first_choice_rate': first_choice_rate,
            'avg_team_size': avg_team_size,
            'constraint_violations': option.constraint_violations
        }
    
    def export_csv_files(self, scheduling_result: SchedulingResult, output_dir: str) -> List[str]:
        """CSV 파일 내보내기 (Gmail/Outlook 호환)"""
        
        output_paths = []
        selected_option = scheduling_result.selected_option or scheduling_result.options[0]
        
        if not selected_option:
            return output_paths
        
        # 메일 머지 데이터 준비
        mail_data = []
        for schedule in selected_option.schedules:
            row_data = {
                '팀명': schedule.team_name,
                '대표자명': schedule.leader_name,
                '이메일': schedule.team.primary_email,
                '연락처': schedule.team.primary_phone,
                '면접날짜': schedule.interview_date,
                '면접시간': schedule.interview_time,
                '면접조': schedule.interview_group,
                '면접장소': schedule.interview_slot.room,
                '줌링크': schedule.interview_slot.zoom_link or "",
                '추가안내사항': schedule.notes or ""
            }
            mail_data.append(row_data)
        
        df = pd.DataFrame(mail_data)
        
        # Gmail용 CSV (UTF-8 BOM)
        timestamp = datetime.now().strftime("%Y%m%d")
        gmail_path = Path(output_dir) / f"gmail_merge_{timestamp}.csv"
        df.to_csv(gmail_path, encoding='utf-8-sig', index=False)
        output_paths.append(str(gmail_path))
        
        # Outlook용 CSV (CP949)
        outlook_path = Path(output_dir) / f"outlook_merge_{timestamp}.csv"
        try:
            df.to_csv(outlook_path, encoding='cp949', index=False)
            output_paths.append(str(outlook_path))
        except UnicodeEncodeError:
            # CP949 인코딩 실패시 UTF-8로 대체
            df.to_csv(outlook_path, encoding='utf-8-sig', index=False)
            output_paths.append(str(outlook_path))
        
        logger.info(f"CSV 파일 내보내기 완료: {len(output_paths)}개 파일")
        return output_paths
    
    def get_workbook_bytes(self) -> bytes:
        """워크북을 바이트 배열로 반환 (스트리밍용)"""
        if not self.wb:
            raise ValueError("워크북이 생성되지 않았습니다.")
        
        output = io.BytesIO()
        self.wb.save(output)
        return output.getvalue()